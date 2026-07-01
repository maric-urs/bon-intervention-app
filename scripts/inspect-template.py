import openpyxl
from pathlib import Path

p = Path(__file__).resolve().parents[1] / "assets" / "bon-intervention-cacem.xlsx"
wb = openpyxl.load_workbook(p)
print("sheets:", wb.sheetnames)
ws = wb["SAISIE"]
for addr in ["B4", "E4", "H4", "B5", "E5", "H5", "B6", "E6", "B9", "E9", "H9", "H21"]:
    print(addr, ws[addr].value)
for r in range(13, 16):
    print("row", r, [ws.cell(r, c).value for c in range(1, 10)])
