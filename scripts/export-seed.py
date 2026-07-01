"""Export fleet and reference data for Next.js seed."""
import json
import re
from pathlib import Path

import openpyxl
import pdfplumber

ROOT = Path(__file__).parent.parent
EXCEL = ROOT.parent / "Bon d'intervention.xlsx"
BUILD = ROOT.parent / "build_bon_v3.py"

import importlib.util
spec = importlib.util.spec_from_file_location("build", BUILD)
b = importlib.util.module_from_spec(spec)
spec.loader.exec_module(b)

def norm_key(s):
    return b.norm_dim_key(s)

def parse_bpu(pdf_path, lot):
    return b.parse_bpu_tires(pdf_path, lot)

def parse_services(pdf_path, lot):
    return b.parse_bpu_services(pdf_path, lot)

wb = openpyxl.load_workbook(EXCEL, data_only=True)
ws = wb[[n for n in wb.sheetnames if "VEHICULE" in n.upper()][0]]

vehicles = []
seen_immat = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row or not row[1]:
        continue
    immat = str(row[1]).strip()
    dim = str(row[5]).strip() if row[5] else ""
    dim = b.resolve_vehicle_dimension(immat, row[2], row[3], dim)
    lot = b.suggest_lot(dim, str(row[3] or ""))
    key = f"{immat}|{str(row[4]).strip()}"
    vehicles.append({
        "immatriculation": immat,
        "marque": str(row[2] or "").strip(),
        "modele": str(row[3] or "").strip(),
        "emplacement": str(row[4] or "").strip(),
        "dimension": dim,
        "lotSuggere": lot,
        "cleTarif": norm_key(dim),
    })

tires = b.merge_tire_tariffs(
    parse_bpu(ROOT.parent / "2025MP061BPU_Lot_1_pneumatiques_CACEM.pdf", "Lot 1")
    + parse_bpu(ROOT.parent / "2025MP061BPU_LOT_2.pdf", "Lot 2")
)
for t in b.SUPPLEMENTAL_TIRES:
    key = f"{t['lot']}|{t['cle']}"
    if not any(f"{x['lot']}|{x['cle']}" == key for x in tires):
        tires.append(t)

services = []
seen = set()
for pdf, lot in [
    (ROOT.parent / "2025MP061BPU_Lot_1_pneumatiques_CACEM.pdf", "Lot 1"),
    (ROOT.parent / "2025MP061BPU_LOT_2.pdf", "Lot 2"),
]:
    for s in parse_services(pdf, lot):
        k = (s["lot"], s["prestation"])
        if k not in seen:
            seen.add(k)
            services.append(s)

centres = [
    {"lot": "Lot 1-2", "nom": "Pneu Cash Schoelcher", "adresse": "Zone de la carrière Ernoult, 97233 Schoelcher",
     "email": "pcschoelcher@citadelle-sa.com", "mobile": "0696 31 97 16", "fixe": "0596 57 25 25", "emailCc": "lbarru@citadelle-sa.com"},
    {"lot": "Lot 1-2", "nom": "Pneu Cash Les Mangles", "adresse": "ZI Les Mangles, 97232 Le Lamentin",
     "email": "pcmangles@citadelle-sa.com", "mobile": "0696 31 97 16", "fixe": "0596 57 25 25", "emailCc": "lbarru@citadelle-sa.com"},
    {"lot": "Lot 1-2", "nom": "Pneu Cash Place d'Armes", "adresse": "CC Place d'armes, 97232 Le Lamentin",
     "email": "pcplacedarmes@citadelle-sa.com", "mobile": "0696 31 97 16", "fixe": "0596 57 25 25", "emailCc": "lbarru@citadelle-sa.com"},
    {"lot": "Lot 1-2", "nom": "Pneu Cash Génipa", "adresse": "ZAC de Génipa, 97224 Ducos",
     "email": "pcgenipa@citadelle-sa.com", "mobile": "0696 31 97 16", "fixe": "0596 57 25 25", "emailCc": "lbarru@citadelle-sa.com"},
    {"lot": "Lot 1-2", "nom": "Pneu Cash Trinité", "adresse": "ZAC du BAC, 97220 Trinité",
     "email": "pctrinite@citadelle-sa.com", "mobile": "0696 31 97 16", "fixe": "0596 57 25 25", "emailCc": "lbarru@citadelle-sa.com"},
    {"lot": "Lot 3", "nom": "SOMAREC Aéroport", "adresse": "Quartier Aéroport - RN5, 97232 Le Lamentin",
     "email": "contact.somarec@gbh.fr", "mobile": "", "fixe": "0596 42 44 24", "emailCc": ""},
]

out = ROOT / "prisma" / "seed-data.json"
out.parent.mkdir(parents=True, exist_ok=True)
out.write_text(json.dumps({
    "marche": {"reference": "25.061", "consultation": "202507011207", "maitre": "CACEM"},
    "centres": centres,
    "vehicles": vehicles,
    "tarifsPneus": tires,
    "tarifsPrestations": services,
}, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"Exported {len(vehicles)} vehicles, {len(tires)} tarifs, {len(centres)} centres -> {out}")
